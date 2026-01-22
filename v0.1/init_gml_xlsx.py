from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import datetime as dt

def autosize(ws, max_col=20, min_width=10, max_width=55):
    for col in range(1, min(ws.max_column, max_col) + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(min_width, min(max_width, max_len + 2))

def header_row(ws, row=1):
    fill = PatternFill("solid", fgColor="F2F2F2")
    for cell in ws[row]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def main(out_path="GML_v0_1.xlsx"):
    wb = Workbook()

    # README
    ws = wb.active
    ws.title = "README"
    ws["A1"] = "GML v0.1 — 变强系统（最小可用版）"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A3"] = "核心：记录行为 → 看到变强 → 产生动力（系统每天≤3分钟）"
    ws["A5"] = "用法：先用 gml_cli.py 记录，再可选 reveal（每日最多3次）"
    ws.column_dimensions["A"].width = 90

    # CONFIG
    ws = wb.create_sheet("CONFIG")
    ws.append(["key", "value", "note"])
    ws.append(["version", "0.1", "file format version"])
    ws.append(["created_at", dt.datetime.now().isoformat(timespec="seconds"), ""])
    ws.append(["player_name", "Alan", "optional"])
    ws.append(["daily_reveal_limit", 3, "max random drops per day"])
    ws.append(["level_xp_base", 200, "xp needed for level 2 is base"])
    ws.append(["level_xp_growth", 1.18, "next level xp multiplier"])
    header_row(ws); autosize(ws)

    # STATE
    ws = wb.create_sheet("STATE")
    ws.append(["key", "value", "note"])
    ws.append(["total_xp", 0, "sum of all xp"])
    ws.append(["level", 1, "cached"])
    ws.append(["title", "新手冒险者", "role/title"])
    ws.append(["streak_days", 0, "consecutive days with any log"])
    ws.append(["last_log_date", "", "YYYY-MM-DD"])
    ws.append(["reveal_count_today", 0, "resets daily"])
    ws.append(["reveal_date", "", "YYYY-MM-DD"])
    for d in ["BODY", "MATH", "CODE", "LIFE", "EXPLORE"]:
        ws.append([f"xp_{d}", 0, "attribute xp"])
    header_row(ws); autosize(ws)

    # TASKS
    ws = wb.create_sheet("TASKS")
    ws.append(["task_id","task_name","domain","xp_per_min","min_minutes","cooldown_days","active","notes"])
    tasks = [
        ("BODY01","徒手训练/有氧","BODY",2,20,0,1,"健康主线"),
        ("MATH01","数学深度学习/习题","MATH",2,30,0,1,"数学主线"),
        ("CODE01","代码练习/项目/刷题","CODE",2,30,0,1,"代码主线"),
        ("LIFE01","做饭/备餐","LIFE",2,15,0,1,"生活质量"),
        ("LIFE02","洗澡+刷牙+基础卫生","LIFE",2,10,1,1,"冷却=1天"),
        ("LIFE03","洗衣/洗内裤/洗袜子","LIFE",2,15,2,1,"冷却=2天"),
        ("LIFE04","吸尘/整理桌面/收纳","LIFE",2,15,3,1,"冷却=3天"),
        ("EXP01","周末探索/美食/绘画等","EXPLORE",2,30,0,1,"探索世界"),
    ]
    for t in tasks:
        ws.append(list(t))
    header_row(ws); autosize(ws, max_col=15)

    # LOG
    ws = wb.create_sheet("LOG")
    ws.append(["timestamp","date","task_id","task_name","domain","minutes","xp","notes"])
    header_row(ws); autosize(ws, max_col=10)
    ws.freeze_panes = "A2"

    # REWARDS
    ws = wb.create_sheet("REWARDS")
    ws.append(["reward_id","reward_name","type","weight","effect","note"])
    rewards = [
        ("R001","下一次XP ×1.2","boost",12,"next_xp_mult:1.2","只影响下一条记录"),
        ("R002","今日允许‘轻松奖励’一次","perk",10,"soft_reward_token:+1","比如喜欢的饮品"),
        ("R003","明天主线减压券","perk",8,"tomorrow_min_minutes:-10","降低启动门槛"),
        ("R004","称号彩蛋：自尊宣言","cosmetic",20,"affirmation:+1","一句短宣言"),
        ("R005","无掉落（继续努力）","none",14,"none","保持随机但不过度刺激"),
    ]
    for r in rewards:
        ws.append(list(r))
    header_row(ws); autosize(ws, max_col=10)

    # INVENTORY
    ws = wb.create_sheet("INVENTORY")
    ws.append(["timestamp","date","reward_id","reward_name","effect","consumed","note"])
    header_row(ws); autosize(ws, max_col=10)
    ws.freeze_panes = "A2"

    wb.save(out_path)
    print(f"Created: {out_path}")

if __name__ == "__main__":
    main()

