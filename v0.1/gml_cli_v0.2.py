#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
GML v0.2 — Weekly Chest (NO reward pool)

Sheets expected in xlsx:
- Tasks:   Task_ID | Name | Weight | Mode(minutes/time/check) | Note
- Log:     Date | Task_ID | Minutes | Meta
- CONFIG:  Key | Value | Note
- STATE:   Key | Value | Note
- Rewards: Week_Start | Week_End | Qualified | Streak_Weeks_After | Chest_Floor | Chest_Roll | Pity_b_no_a | Pity_a_no_s | Note
"""

from __future__ import annotations
import argparse
import datetime as dt
import os
import sys
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook


S_TASKS = "TASKS"
S_LOG = "LOG"
S_CONFIG = "CONFIG"
S_STATE = "STATE"
S_REWARDS = "REWARDS"


def die(msg: str, code: int = 1) -> None:
    print(f"[ERROR] {msg}", file=sys.stderr)
    raise SystemExit(code)


def parse_date(s: Optional[str]) -> dt.date:
    if not s:
        return dt.date.today()
    try:
        return dt.date.fromisoformat(s)
    except Exception:
        die(f"Bad --date '{s}'. Use YYYY-MM-DD.")


def parse_hhmm(s: str) -> Optional[int]:
    s = (s or "").strip()
    if not s:
        return None
    try:
        hh, mm = s.split(":")
        hh_i, mm_i = int(hh), int(mm)
        if not (0 <= hh_i <= 23 and 0 <= mm_i <= 59):
            return None
        return hh_i * 60 + mm_i
    except Exception:
        return None


def in_window(value_min: int, target_min: int, tol_min: int) -> bool:
    return (target_min - tol_min) <= value_min <= (target_min + tol_min)


def load_wb(path: str) -> Workbook:
    if not os.path.exists(path):
        raise FileNotFoundError(f"Excel file not found: {path}")
    wb = load_workbook(path)
    for sh in [S_TASKS, S_LOG, S_CONFIG, S_STATE, S_REWARDS]:
        if sh not in wb.sheetnames:
            raise KeyError(f"Worksheet '{sh}' does not exist.")
    return wb


def read_kv(wb: Workbook, sheet: str) -> Dict[str, str]:
    ws = wb[sheet]
    out: Dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        k = str(row[0]).strip()
        v = "" if row[1] is None else str(row[1]).strip()
        out[k] = v
    return out


def write_kv(wb: Workbook, sheet: str, updates: Dict[str, str]) -> None:
    ws = wb[sheet]
    key_row: Dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        k = ws.cell(row=r, column=1).value
        if k:
            key_row[str(k).strip()] = r
    for k, v in updates.items():
        if k in key_row:
            ws.cell(row=key_row[k], column=2).value = v
        else:
            ws.append([k, v, ""])


@dataclass
class Task:
    task_id: str
    name: str
    weight: int
    mode: str  # minutes | time | check
    note: str

def read_tasks(wb: Workbook) -> List[Task]:
    ws = wb[S_TASKS]
    tasks: List[Task] = []

    def to_int(x, default=None):
        try:
            if x is None:
                return default
            if isinstance(x, str):
                x = x.strip()
                if x == "":
                    return default
            return int(x)
        except (ValueError, TypeError):
            return default

    def to_str(x, default=""):
        if x is None:
            return default
        return str(x).strip()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        tid = to_str(row[0]).upper()
        name = to_str(row[1])

        # 兼容两种结构：
        # A: [Task_ID, Name, Weight, Mode, Note]
        # B: [Task_ID, Name, Group, Weight, Mode, Note]
        weight = to_int(row[2], default=None)

        if weight is None:
            # 说明 row[2] 不是数字（例如 'BODY'），走结构 B
            weight = to_int(row[3], default=0)
            mode = to_str(row[4], default="minutes").lower() if len(row) > 4 else "minutes"
            note = to_str(row[5], default="") if len(row) > 5 else ""
        else:
            # 结构 A
            mode = to_str(row[3], default="minutes").lower() if len(row) > 3 else "minutes"
            note = to_str(row[4], default="") if len(row) > 4 else ""

        tasks.append(Task(tid, name, weight, mode, note))

    return tasks

def append_log(wb: Workbook, date: dt.date, task_id: str, minutes: int, meta: str) -> None:
    wb[S_LOG].append([date.isoformat(), task_id.upper(), int(minutes), meta])


def iter_log(wb: Workbook) -> List[Tuple[dt.date, str, int, str]]:
    ws = wb[S_LOG]
    out: List[Tuple[dt.date, str, int, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not row[1]:
            continue
        try:
            d = dt.date.fromisoformat(str(row[0]))
        except Exception:
            continue
        tid = str(row[1]).strip().upper()
        mins = int(row[2] or 0)
        meta = "" if row[3] is None else str(row[3]).strip()
        out.append((d, tid, mins, meta))
    return out


# ===== Weekly window: last full week ending on last Sunday strictly before today =====
def last_full_week(today: dt.date) -> Tuple[dt.date, dt.date]:
    # weekday: Mon=0 ... Sun=6
    offset = today.weekday() + 1  # Mon->1 ... Sun->7
    week_end = today - dt.timedelta(days=offset)  # previous Sunday (strictly before today)
    week_start = week_end - dt.timedelta(days=6)
    return week_start, week_end


def parse_streak_map(s: str) -> List[Tuple[int, str]]:
    pairs: List[Tuple[int, str]] = []
    for part in (s or "").split(","):
        part = part.strip()
        if not part:
            continue
        left, right = part.split(":")
        pairs.append((int(left.strip()), right.strip().upper()))
    pairs.sort(key=lambda x: x[0])
    return pairs


def chest_floor_from_streak(streak: int, streak_map: List[Tuple[int, str]]) -> str:
    floor = "D"
    for n, tier in streak_map:
        if streak >= n:
            floor = tier
    return floor


def roll_chest(floor: str, pity_b_no_a: int, pity_a_no_s: int,
              pity_B_to_A: int, pity_A_to_S: int) -> Tuple[str, int, int]:
    """
    NO reward pool, only returns tier.
    Default distributions (you can tweak later).
    """
    import random
    floor = floor.upper()

    # pity triggers
    if floor == "B" and pity_b_no_a >= pity_B_to_A:
        return "A", 0, pity_a_no_s
    if floor == "A" and pity_a_no_s >= pity_A_to_S:
        return "S", pity_b_no_a, 0

    r = random.random()
    if floor == "D":
        tier = "D" if r < 0.80 else ("C" if r < 0.98 else "B")
    elif floor == "C":
        tier = "C" if r < 0.70 else ("B" if r < 0.95 else "A")
    elif floor == "B":
        tier = "B" if r < 0.70 else ("A" if r < 0.95 else "S")
    elif floor == "A":
        tier = "A" if r < 0.85 else "S"
    else:
        tier = "S"

    # update pity
    if floor == "B":
        pity_b_no_a = 0 if tier in ("A", "S") else pity_b_no_a + 1
    if floor == "A":
        pity_a_no_s = 0 if tier == "S" else pity_a_no_s + 1
    return tier, pity_b_no_a, pity_a_no_s


def already_opened(wb: Workbook, week_start: dt.date, week_end: dt.date) -> bool:
    ws = wb[S_REWARDS]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not row[1]:
            continue
        if str(row[0]).strip() == week_start.isoformat() and str(row[1]).strip() == week_end.isoformat():
            return True
    return False


def week_stats(wb: Workbook, week_start: dt.date, week_end: dt.date, cfg: Dict[str, str]) -> Dict[str, int]:
    logs = iter_log(wb)
    deep_thr = int(cfg.get("deep_minutes_threshold", "30") or 30)

    days = [week_start + dt.timedelta(days=i) for i in range(7)]
    food_done = {d: set() for d in days}
    sleep_ok = {d: {"SLEEP_PM": False, "WAKE_AM": False} for d in days}

    bed_target = parse_hhmm(cfg.get("sleep_bed_target", "22:10")) or (22 * 60 + 10)
    bed_tol = int(cfg.get("sleep_bed_tol_min", "20") or 20)
    wake_target = parse_hhmm(cfg.get("wake_target", "07:30")) or (7 * 60 + 30)
    wake_tol = int(cfg.get("wake_tol_min", "10") or 10)

    deep_math = 0
    deep_code = 0

    for d, tid, mins, meta in logs:
        if d < week_start or d > week_end:
            continue

        if tid == "MATH01" and mins >= deep_thr:
            deep_math += 1
        if tid == "CODE01" and mins >= deep_thr:
            deep_code += 1

        if tid in ("FOOD_B", "FOOD_L", "FOOD_D"):
            food_done[d].add(tid)

        if tid in ("SLEEP_PM", "WAKE_AM"):
            tstr = meta
            if "t=" in meta:
                tstr = meta.split("t=", 1)[1].strip()
            val = parse_hhmm(tstr)
            if val is None:
                continue
            if tid == "SLEEP_PM" and in_window(val, bed_target, bed_tol):
                sleep_ok[d]["SLEEP_PM"] = True
            if tid == "WAKE_AM" and in_window(val, wake_target, wake_tol):
                sleep_ok[d]["WAKE_AM"] = True

    pass_days = 0
    for d in days:
        food_ok = food_done[d] >= {"FOOD_B", "FOOD_L", "FOOD_D"}
        day_sleep_ok = sleep_ok[d]["SLEEP_PM"] and sleep_ok[d]["WAKE_AM"]
        if food_ok and day_sleep_ok:
            pass_days += 1

    deep_total = deep_math + deep_code
    return {"pass_days": pass_days, "deep_total": deep_total, "deep_math": deep_math, "deep_code": deep_code}


def reveal_weekly(wb: Workbook, today: dt.date) -> None:
    cfg = read_kv(wb, S_CONFIG)
    st = read_kv(wb, S_STATE)

    week_start, week_end = last_full_week(today)
    print("\n--- Weekly chest (last full week) ---")
    print(f"Target week: {week_start.isoformat()} ~ {week_end.isoformat()}")

    if already_opened(wb, week_start, week_end):
        print("Already opened. (See Rewards sheet)")
        return

    stats = week_stats(wb, week_start, week_end, cfg)

    req_pass = int(cfg.get("weekly_pass_days_required", "5") or 5)
    req_deep = int(cfg.get("weekly_deep_sessions_required", "5") or 5)
    req_math = int(cfg.get("weekly_deep_math_required", "2") or 2)
    req_code = int(cfg.get("weekly_deep_code_required", "2") or 2)

    qualified = (
        stats["pass_days"] >= req_pass and
        stats["deep_total"] >= req_deep and
        stats["deep_math"] >= req_math and
        stats["deep_code"] >= req_code
    )

    print(f"Pass days: {stats['pass_days']}/7 (need >= {req_pass})")
    print(f"Deep: total={stats['deep_total']} (MATH={stats['deep_math']}, CODE={stats['deep_code']}) "
          f"(need total>={req_deep}, MATH>={req_math}, CODE>={req_code})")

    streak = int(st.get("streak_weeks", "0") or 0)
    pity_b = int(st.get("pity_b_no_a", "0") or 0)
    pity_a = int(st.get("pity_a_no_s", "0") or 0)

    ws = wb[S_REWARDS]

    if not qualified:
        print("NOT qualified → no chest, streak resets to 0.")
        ws.append([week_start.isoformat(), week_end.isoformat(), "N", 0, "-", "-", pity_b, pity_a, "not qualified"])
        write_kv(wb, S_STATE, {"streak_weeks": "0"})
        return

    # qualified
    streak_after = streak + 1
    streak_map = parse_streak_map(cfg.get("chest_streak_map", "1:D,2:C,4:B,8:A,16:S"))
    floor = chest_floor_from_streak(streak_after, streak_map)

    pity_B_to_A = int(cfg.get("pity_B_to_A", "4") or 4)
    pity_A_to_S = int(cfg.get("pity_A_to_S", "8") or 8)

    roll, pity_b2, pity_a2 = roll_chest(floor, pity_b, pity_a, pity_B_to_A, pity_A_to_S)

    print(f"Qualified ✅  | Streak: {streak} -> {streak_after}")
    print(f"Chest floor: {floor} | Roll: {roll}")

    ws.append([week_start.isoformat(), week_end.isoformat(), "Y", streak_after, floor, roll, pity_b2, pity_a2, ""])
    write_kv(wb, S_STATE, {
        "streak_weeks": str(streak_after),
        "pity_b_no_a": str(pity_b2),
        "pity_a_no_s": str(pity_a2),
    })


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", default="GML_v0_1.xlsx")
    ap.add_argument("--date", default=None)
    ap.add_argument("--reveal", action="store_true")
    args = ap.parse_args()

    date = parse_date(args.date)
    wb = load_wb(args.file)
    tasks = read_tasks(wb)

    print(f"\nGML v0.2 — {date.isoformat()}")
    print("\nAvailable tasks:")
    for t in tasks:
        print(f"  {t.task_id:<8} {t.name} {t.weight}")

    s = input("\nPick tasks you completed today (comma separated Task_ID), or Enter to skip.\nTask_IDs: ").strip()
    if s:
        selected = [x.strip().upper() for x in s.split(",") if x.strip()]
        now = dt.datetime.now()
        for tid in selected:
            t = next((x for x in tasks if x.task_id == tid), None)
            if not t:
                print(f"[WARN] Unknown Task_ID: {tid}")
                continue

            if t.mode == "minutes":
                raw = input(f"Minutes for {tid} (default 0): ").strip()
                mins = int(raw) if raw else 0
                append_log(wb, date, tid, mins, "")
            elif t.mode == "time":
                default = now.strftime("%H:%M")
                raw = input(f"Time for {tid} HH:MM (blank=now {default}): ").strip()
                tt = raw if raw else default
                append_log(wb, date, tid, 0, f"t={tt}")
            elif t.mode == "check":
                raw = input(f"{tid} done? (Y/n, default Y): ").strip().lower()
                ok = (raw != "n")
                append_log(wb, date, tid, 0, "y" if ok else "n")
            else:
                raw = input(f"Minutes for {tid} (default 0): ").strip()
                mins = int(raw) if raw else 0
                append_log(wb, date, tid, mins, "")

        wb.save(args.file)
        print("[OK] Logged.")

    if args.reveal:
        wb = load_wb(args.file)
        reveal_weekly(wb, date)
        wb.save(args.file)
        print("[OK] Reveal saved.")


if __name__ == "__main__":
    main()

