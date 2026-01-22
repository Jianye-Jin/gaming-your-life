#!/usr/bin/env python3
"""
Gaming My Life (GML) — v0.1 CLI logger

What this does:
- Append completed tasks into Daily_Log (event log)
- Compute today's Daily Pass gate: BODY>=1 and MAIN>=1 and HOME>=1
- If pass: mark today's Chest_Queue row as Eligible=1 and (optionally) Revealed=1 with timestamp

Usage:
  python gml_cli.py --file GML_v0_1.xlsx
"""
from __future__ import annotations
import argparse
import datetime as dt
from pathlib import Path

import openpyxl

def today_str() -> str:
    return dt.date.today().isoformat()

def load_wb(path: Path) -> openpyxl.Workbook:
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")
    return openpyxl.load_workbook(path)

def read_tasks(wb: openpyxl.Workbook):
    ws = wb["TASKS"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    tasks = []
    for r in rows:
        if not r[0]:
            continue
        tasks.append({
            "Task_ID": str(r[0]).strip(),
            "Category": str(r[1]).strip(),
            "Cadence": str(r[2]).strip(),
            "Task_Name": str(r[3]).strip(),
            "Default_Minutes": r[5],
        })
    return tasks

def append_log(wb: openpyxl.Workbook, date: str, task_id: str, category: str, minutes: int, notes: str=""):
    ws = wb["Daily_Log"]
    # Find next row (append)
    next_row = ws.max_row + 1
    ts = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    entry_id = next_row - 1  # simple incremental id (header=1)
    ws.cell(next_row, 1, entry_id)
    ws.cell(next_row, 2, date)
    ws.cell(next_row, 3, task_id)
    ws.cell(next_row, 4, category)
    ws.cell(next_row, 5, minutes)
    ws.cell(next_row, 6, "")  # quality optional
    ws.cell(next_row, 7, notes)
    ws.cell(next_row, 8, "")  # evidence link optional
    ws.cell(next_row, 9, ts)

def counts_for_date(wb: openpyxl.Workbook, date: str):
    ws = wb["Daily_Log"]
    # Columns: B Date, D Category
    counts = {"BODY":0,"MAIN":0,"HOME":0,"EXP":0}
    total_minutes = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        if str(row[1]).strip() != date:
            continue
        cat = (str(row[3]).strip() if row[3] else "")
        if cat in counts:
            counts[cat] += 1
        try:
            total_minutes += int(row[4] or 0)
        except Exception:
            pass
    return counts, total_minutes

def mark_chest(wb: openpyxl.Workbook, date: str, reveal: bool):
    ws = wb["Chest_Queue"]
    # find matching date in col A
    for r in range(2, ws.max_row+1):
        if str(ws.cell(r,1).value).strip() == date:
            ws.cell(r,2).value = 1  # eligible
            if reveal:
                ws.cell(r,5).value = 1
                ws.cell(r,6).value = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            return True
    return False

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", type=str, default="GML_v0_1.xlsx")
    ap.add_argument("--date", type=str, default=today_str())
    ap.add_argument("--reveal", action="store_true", help="also reveal (open) today's chest if Daily Pass")
    args = ap.parse_args()

    path = Path(args.file).expanduser().resolve()
    wb = load_wb(path)
    tasks = read_tasks(wb)

    # Quick menu
    print(f"\nGML v0.1 — {args.date}\n")
    print("Pick tasks you completed today (comma separated Task_ID), or press Enter to skip logging.\n")
    print("Available tasks:")
    for t in tasks:
        print(f"  {t['Task_ID']:6s} [{t['Category']}] {t['Task_Name']}")
    raw = input("\nTask_IDs: ").strip()
    if raw:
        ids = [x.strip().upper() for x in raw.split(",") if x.strip()]
        for tid in ids:
            match = next((t for t in tasks if t["Task_ID"] == tid), None)
            if not match:
                print(f"  ! Unknown Task_ID: {tid} (skipped)")
                continue
            mins_raw = input(f"  Minutes for {tid} (default {match['Default_Minutes']}): ").strip()
            minutes = int(mins_raw) if mins_raw else int(match["Default_Minutes"] or 0)
            notes = input("  Notes (optional): ").strip()
            append_log(wb, args.date, tid, match["Category"], minutes, notes)
            print("  ✓ logged")
    counts, total_minutes = counts_for_date(wb, args.date)
    daily_pass = (counts["BODY"]>=1 and counts["MAIN"]>=1 and counts["HOME"]>=1)

    print("\nToday summary:")
    print(f"  BODY={counts['BODY']}  MAIN={counts['MAIN']}  HOME={counts['HOME']}  EXP={counts['EXP']}  minutes={total_minutes}")
    print(f"  Daily Pass: {'YES' if daily_pass else 'NO'} (needs BODY>=1 & MAIN>=1 & HOME>=1)")

    if daily_pass:
        ok = mark_chest(wb, args.date, reveal=args.reveal)
        if ok:
            print("  Chest: eligible marked" + (" + revealed" if args.reveal else ""))
        else:
            print("  Chest: could not find today's row in Chest_Queue")

    wb.save(path)
    print(f"\nSaved: {path}\n")

if __name__ == "__main__":
    main()
