import os
import csv
from openpyxl import load_workbook

XLSX = "GML_v0_1.xlsx"
OUTDIR = "csv_export"

def sheet_to_csv(ws, path):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            w.writerow(["" if v is None else v for v in row])

def main():
    os.makedirs(OUTDIR, exist_ok=True)
    wb = load_workbook(XLSX)
    for name in wb.sheetnames:
        ws = wb[name]
        sheet_to_csv(ws, os.path.join(OUTDIR, f"{name}.csv"))
    print(f"Exported CSV to: {OUTDIR}/")

if __name__ == "__main__":
    main()
