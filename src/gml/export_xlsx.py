from __future__ import annotations

from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from gml import db


def _style_header(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"


def _autosize(ws, max_width: int = 50) -> None:
    # simple autosize (good enough)
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        width = min(max_len + 2, max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, width)


def export_xlsx(
    db_path: Path,
    out_path: Path,
    since: Optional[str] = None,
    until: Optional[str] = None,
) -> Path:
    """
    Export TASKS / LOG / CHESTS from SQLite into an .xlsx workbook.
    since/until: filter by date (YYYY-MM-DD), inclusive.
    """
    db_path = db_path.expanduser().resolve()
    out_path = out_path.expanduser().resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)

    conn = db.connect(db_path)
    try:
        # TASKS
        tasks = conn.execute(
            """
            SELECT id, name, domain, cadence, default_minutes, default_xp, active, created_at
            FROM tasks
            ORDER BY domain, id
            """
        ).fetchall()

        # LOG (joined)
        where = []
        params = []
        if since:
            where.append("l.date >= ?")
            params.append(since)
        if until:
            where.append("l.date <= ?")
            params.append(until)
        where_sql = ("WHERE " + " AND ".join(where)) if where else ""

        logs = conn.execute(
            f"""
            SELECT
              l.ts       AS timestamp,
              l.date     AS date,
              l.task_id  AS task_id,
              t.name     AS task_name,
              t.domain   AS domain,
              l.minutes  AS minutes,
              l.xp       AS xp,
              l.notes    AS notes
            FROM logs l
            JOIN tasks t ON t.id = l.task_id
            {where_sql}
            ORDER BY l.date, l.ts, l.id
            """,
            params,
        ).fetchall()

        # CHESTS
        c_where = []
        c_params = []
        if since:
            c_where.append("date >= ?")
            c_params.append(since)
        if until:
            c_where.append("date <= ?")
            c_params.append(until)
        c_where_sql = ("WHERE " + " AND ".join(c_where)) if c_where else ""

        chests = conn.execute(
            f"""
            SELECT date, eligible, revealed, revealed_ts
            FROM chests
            {c_where_sql}
            ORDER BY date
            """,
            c_params,
        ).fetchall()

    finally:
        conn.close()

    wb = openpyxl.Workbook()

    # TASKS sheet
    ws = wb.active
    ws.title = "TASKS"
    ws.append(["id", "name", "domain", "cadence", "default_minutes", "default_xp", "active", "created_at"])
    for r in tasks:
        ws.append([r["id"], r["name"], r["domain"], r["cadence"], r["default_minutes"], r["default_xp"], r["active"], r["created_at"]])
    _style_header(ws)
    _autosize(ws)

    # LOG sheet (align with your init_gml_xlsx.py header)
    ws = wb.create_sheet("LOG")
    ws.append(["timestamp", "date", "task_id", "task_name", "domain", "minutes", "xp", "notes"])
    for r in logs:
        ws.append([r["timestamp"], r["date"], r["task_id"], r["task_name"], r["domain"], r["minutes"], r["xp"], r["notes"]])
    _style_header(ws)
    _autosize(ws)

    # CHESTS sheet
    ws = wb.create_sheet("CHESTS")
    ws.append(["date", "eligible", "revealed", "revealed_ts"])
    for r in chests:
        ws.append([r["date"], r["eligible"], r["revealed"], r["revealed_ts"]])
    _style_header(ws)
    _autosize(ws)

    wb.save(str(out_path))
    return out_path

